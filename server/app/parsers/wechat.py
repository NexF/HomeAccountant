"""微信账单 xlsx 解析器

支持格式：微信「账单 → 导出账单」功能导出的 xlsx 文件。
文件结构：
  Row 0: "微信支付账单明细"（标识行）
  Row 1-14: 元信息/汇总/注释
  Row 15: 分隔线
  Row 16: 表头（交易时间|交易类型|交易对方|商品|收/支|金额(元)|支付方式|当前状态|交易单号|商户单号|备注）
  Row 17+: 数据行
"""

import io
from datetime import datetime

from openpyxl import load_workbook


# 成功状态关键词
SUCCESS_KEYWORDS = ("成功", "已存入", "已收钱", "已转账", "已到账")


def parse_wechat_xlsx(content: bytes) -> list[dict]:
    """解析微信账单 xlsx 文件内容，返回标准化行列表。

    Args:
        content: xlsx 文件的二进制内容

    Returns:
        list[dict]: 每个 dict 包含 date, description, amount, direction,
                    payment_method, external_id, is_duplicate(默认 False)

    Raises:
        ValueError: 非微信账单格式或无有效数据
    """
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("无法读取工作表")

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("文件为空")

    # 1. 校验标识行
    first_cell = str(rows[0][0]).strip() if rows[0][0] else ""
    if "微信支付账单明细" not in first_cell:
        raise ValueError("未找到「微信支付账单明细」标识，请确认文件为微信导出的账单")

    # 2. 定位表头行（找到第一列为"交易时间"的行）
    header_row_idx = None
    for i, row in enumerate(rows):
        cell = str(row[0]).strip() if row[0] else ""
        if cell == "交易时间":
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("未找到表头行（交易时间列）")

    # 3. 解析数据行
    result = []
    for row_idx in range(header_row_idx + 1, len(rows)):
        row = rows[row_idx]
        if not row or not row[0]:
            continue

        # 列映射
        raw_time = str(row[0]).strip()       # 交易时间
        raw_type = str(row[1]).strip() if row[1] else ""  # 交易类型（暂不使用）
        raw_party = str(row[2]).strip() if row[2] else ""  # 交易对方
        raw_goods = str(row[3]).strip() if row[3] else ""  # 商品
        raw_direction = str(row[4]).strip() if row[4] else ""  # 收/支
        raw_amount = str(row[5]).strip() if row[5] else "0"  # 金额(元)
        raw_payment = str(row[6]).strip() if row[6] else ""  # 支付方式
        raw_status = str(row[7]).strip() if row[7] else ""  # 当前状态
        raw_txn_id = str(row[8]).strip() if row[8] else ""  # 交易单号
        # row[9]: 商户单号（暂不使用）
        # row[10]: 备注（暂不使用）

        # 跳过非成功状态
        if not any(kw in raw_status for kw in SUCCESS_KEYWORDS):
            continue

        # 金额解析：去掉 ¥ 前缀，统一正数
        amount_str = raw_amount.replace("¥", "").replace(",", "").strip()
        try:
            amount = abs(float(amount_str))
        except ValueError:
            continue  # 跳过无法解析的行

        if amount == 0:
            continue

        # 描述生成
        if raw_goods and raw_goods != "/":
            description = f"{raw_party} - {raw_goods}"
        else:
            description = raw_party

        # external_id
        external_id = f"wechat_{raw_txn_id.strip()}"

        # 日期解析
        try:
            dt = datetime.strptime(raw_time, "%Y-%m-%d %H:%M:%S")
            date_str = dt.strftime("%Y-%m-%d")
        except ValueError:
            date_str = raw_time[:10]  # fallback

        result.append({
            "date": date_str,
            "description": description,
            "amount": amount,
            "direction": raw_direction,  # 保持原始值：支出 / 收入 / 中性交易
            "payment_method": raw_payment,
            "external_id": external_id,
            "is_duplicate": False,
        })

    return result
