"""微信账单 xlsx 解析器测试

覆盖场景：
- 正常 xlsx 解析（标识行、表头定位、数据提取）
- 金额去 ¥ 前缀、统一正数
- 中性交易正常解析
- 非成功状态行跳过
- 描述生成（商品为 `/` 时只取交易对方）
- external_id 格式正确
- 非微信文件抛 ValueError
- 空文件抛 ValueError
"""

import io

import pytest
from openpyxl import Workbook

from app.parsers.wechat import parse_wechat_xlsx, SUCCESS_KEYWORDS


# ──────────── 辅助函数 ────────────


def _build_wechat_xlsx(
    data_rows: list[list],
    header_row: list | None = None,
    identifier: str = "微信支付账单明细",
    meta_rows: int = 15,
) -> bytes:
    """构造一个模拟微信账单 xlsx 的二进制内容。

    Args:
        data_rows: 数据行列表，每行 11 列
        header_row: 表头行，默认使用标准表头
        identifier: 标识行内容
        meta_rows: 标识行到表头之间的元信息行数（包含分隔行）
    """
    wb = Workbook()
    ws = wb.active

    # Row 1: 标识行
    ws.append([identifier])

    # Row 2 ~ meta_rows: 元信息占位
    for _ in range(meta_rows - 1):
        ws.append([""])

    # 分隔行
    ws.append(["----------------------"])

    # 表头行
    if header_row is None:
        header_row = [
            "交易时间", "交易类型", "交易对方", "商品", "收/支",
            "金额(元)", "支付方式", "当前状态", "交易单号", "商户单号", "备注",
        ]
    ws.append(header_row)

    # 数据行
    for row in data_rows:
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_row(
    time="2026-02-24 12:00:00",
    tx_type="商户消费",
    party="蔡林记",
    goods="热干面",
    direction="支出",
    amount="¥25.00",
    payment="招商银行储蓄卡(3717)",
    status="支付成功",
    txn_id="4200002984202602246232095463",
    merchant_id="123456",
    note="",
):
    """构造单条数据行"""
    return [time, tx_type, party, goods, direction, amount, payment, status, txn_id, merchant_id, note]


# ──────────── 正常解析 ────────────


class TestNormalParsing:

    def test_basic_parsing(self):
        """正常 xlsx 解析：标识行、表头定位、数据提取"""
        rows = [
            _make_row(party="蔡林记", goods="热干面", direction="支出", amount="¥25.00"),
            _make_row(party="Estela", goods="/", direction="收入", amount="¥9.99",
                      txn_id="1000031001000602246315363245890", status="已收钱"),
        ]
        content = _build_wechat_xlsx(rows)
        result = parse_wechat_xlsx(content)

        assert len(result) == 2
        assert result[0]["date"] == "2026-02-24"
        assert result[0]["direction"] == "支出"
        assert result[1]["direction"] == "收入"

    def test_parse_real_file(self):
        """使用真实微信账单文件验证解析"""
        import os
        real_file = os.path.join(
            os.path.dirname(__file__), "..", "..",
            "docs", "v0.4.1",
            "微信支付账单流水文件(20260217-20260224)_20260224173400.xlsx",
        )
        if not os.path.exists(real_file):
            pytest.skip("真实微信账单文件不存在")

        with open(real_file, "rb") as f:
            content = f.read()

        result = parse_wechat_xlsx(content)
        assert len(result) > 0

        # 验证基本结构
        for row in result:
            assert "date" in row
            assert "description" in row
            assert "amount" in row
            assert "direction" in row
            assert "payment_method" in row
            assert "external_id" in row
            assert "is_duplicate" in row

    def test_all_fields_present(self):
        """解析结果包含所有必须字段"""
        content = _build_wechat_xlsx([_make_row()])
        result = parse_wechat_xlsx(content)

        assert len(result) == 1
        row = result[0]
        assert set(row.keys()) == {
            "date", "description", "amount", "direction",
            "payment_method", "external_id", "is_duplicate",
        }

    def test_is_duplicate_defaults_false(self):
        """is_duplicate 默认为 False"""
        content = _build_wechat_xlsx([_make_row()])
        result = parse_wechat_xlsx(content)
        assert result[0]["is_duplicate"] is False


# ──────────── 金额解析 ────────────


class TestAmountParsing:

    def test_amount_strips_yen_prefix(self):
        """金额去掉 ¥ 前缀"""
        content = _build_wechat_xlsx([_make_row(amount="¥25.00")])
        result = parse_wechat_xlsx(content)
        assert result[0]["amount"] == 25.00

    def test_amount_always_positive(self):
        """金额统一为正数，即使原始值带负号"""
        content = _build_wechat_xlsx([_make_row(amount="¥-25.00")])
        result = parse_wechat_xlsx(content)
        assert result[0]["amount"] == 25.00
        assert result[0]["amount"] > 0

    def test_amount_with_comma(self):
        """金额含千分位逗号"""
        content = _build_wechat_xlsx([_make_row(amount="¥1,234.56")])
        result = parse_wechat_xlsx(content)
        assert result[0]["amount"] == 1234.56

    def test_amount_without_yen_prefix(self):
        """金额不带 ¥ 前缀也能解析"""
        content = _build_wechat_xlsx([_make_row(amount="25.00")])
        result = parse_wechat_xlsx(content)
        assert result[0]["amount"] == 25.00

    def test_zero_amount_skipped(self):
        """金额为 0 的行被跳过"""
        content = _build_wechat_xlsx([_make_row(amount="¥0.00")])
        result = parse_wechat_xlsx(content)
        assert len(result) == 0

    def test_invalid_amount_skipped(self):
        """无法解析的金额行被跳过"""
        content = _build_wechat_xlsx([_make_row(amount="N/A")])
        result = parse_wechat_xlsx(content)
        assert len(result) == 0


# ──────────── 中性交易 ────────────


class TestNeutralTransaction:

    def test_neutral_transaction_parsed(self):
        """中性交易正常解析，direction 保持原值"""
        content = _build_wechat_xlsx([
            _make_row(
                party="零钱通", goods="/", direction="中性交易",
                amount="¥100.00", status="已存入",
                txn_id="9900001234567890",
            ),
        ])
        result = parse_wechat_xlsx(content)
        assert len(result) == 1
        assert result[0]["direction"] == "中性交易"
        assert result[0]["amount"] == 100.00

    def test_neutral_with_income_expense_mixed(self):
        """中性交易和收支混合解析"""
        content = _build_wechat_xlsx([
            _make_row(direction="支出", amount="¥25.00", txn_id="001"),
            _make_row(direction="中性交易", amount="¥100.00", status="已存入", txn_id="002"),
            _make_row(direction="收入", amount="¥9.99", status="已收钱", txn_id="003"),
        ])
        result = parse_wechat_xlsx(content)
        assert len(result) == 3
        assert result[0]["direction"] == "支出"
        assert result[1]["direction"] == "中性交易"
        assert result[2]["direction"] == "收入"


# ──────────── 状态过滤 ────────────


class TestStatusFiltering:

    @pytest.mark.parametrize("status_text", [
        "支付成功", "已存入零钱", "已收钱", "已转账", "已到账",
    ])
    def test_success_status_included(self, status_text):
        """包含成功关键词的行被保留"""
        content = _build_wechat_xlsx([_make_row(status=status_text)])
        result = parse_wechat_xlsx(content)
        assert len(result) == 1

    @pytest.mark.parametrize("status_text", [
        "已退款", "对方已退还", "已关闭", "待支付", "",
    ])
    def test_non_success_status_skipped(self, status_text):
        """非成功状态行被跳过"""
        content = _build_wechat_xlsx([_make_row(status=status_text)])
        result = parse_wechat_xlsx(content)
        assert len(result) == 0

    def test_mixed_status_filtering(self):
        """混合状态：只保留成功行"""
        content = _build_wechat_xlsx([
            _make_row(status="支付成功", txn_id="001"),
            _make_row(status="已退款", txn_id="002"),
            _make_row(status="已收钱", txn_id="003"),
        ])
        result = parse_wechat_xlsx(content)
        assert len(result) == 2
        assert result[0]["external_id"] == "wechat_001"
        assert result[1]["external_id"] == "wechat_003"


# ──────────── 描述生成 ────────────


class TestDescriptionGeneration:

    def test_description_with_goods(self):
        """有商品时：交易对方 - 商品"""
        content = _build_wechat_xlsx([_make_row(party="蔡林记", goods="热干面")])
        result = parse_wechat_xlsx(content)
        assert result[0]["description"] == "蔡林记 - 热干面"

    def test_description_goods_is_slash(self):
        """商品为 / 时只取交易对方"""
        content = _build_wechat_xlsx([_make_row(party="Estela", goods="/")])
        result = parse_wechat_xlsx(content)
        assert result[0]["description"] == "Estela"

    def test_description_goods_empty(self):
        """商品为空时只取交易对方"""
        content = _build_wechat_xlsx([_make_row(party="Estela", goods="")])
        result = parse_wechat_xlsx(content)
        assert result[0]["description"] == "Estela"

    def test_description_with_long_goods(self):
        """完整的商品名"""
        content = _build_wechat_xlsx([
            _make_row(party="武汉蔡林记餐饮管理有限公司", goods="蔡林记奥山世纪城店"),
        ])
        result = parse_wechat_xlsx(content)
        assert result[0]["description"] == "武汉蔡林记餐饮管理有限公司 - 蔡林记奥山世纪城店"


# ──────────── external_id ────────────


class TestExternalId:

    def test_external_id_format(self):
        """external_id 格式为 wechat_{交易单号}"""
        content = _build_wechat_xlsx([
            _make_row(txn_id="4200002984202602246232095463"),
        ])
        result = parse_wechat_xlsx(content)
        assert result[0]["external_id"] == "wechat_4200002984202602246232095463"

    def test_external_id_strips_whitespace(self):
        """交易单号的空白被去除"""
        content = _build_wechat_xlsx([
            _make_row(txn_id="  4200002984  "),
        ])
        result = parse_wechat_xlsx(content)
        assert result[0]["external_id"] == "wechat_4200002984"

    def test_external_ids_unique(self):
        """不同行有不同的 external_id"""
        content = _build_wechat_xlsx([
            _make_row(txn_id="001"),
            _make_row(txn_id="002"),
        ])
        result = parse_wechat_xlsx(content)
        assert result[0]["external_id"] != result[1]["external_id"]


# ──────────── 日期解析 ────────────


class TestDateParsing:

    def test_standard_datetime_format(self):
        """标准日期时间格式 YYYY-MM-DD HH:MM:SS"""
        content = _build_wechat_xlsx([_make_row(time="2026-02-24 17:30:00")])
        result = parse_wechat_xlsx(content)
        assert result[0]["date"] == "2026-02-24"

    def test_date_only_fallback(self):
        """非标准格式时截取前10字符作为 fallback"""
        content = _build_wechat_xlsx([_make_row(time="2026-02-24T17:30:00")])
        result = parse_wechat_xlsx(content)
        assert result[0]["date"] == "2026-02-24"


# ──────────── 错误处理 ────────────


class TestErrorHandling:

    def test_non_xlsx_raises_error(self):
        """非 xlsx 文件抛异常"""
        with pytest.raises(Exception):
            parse_wechat_xlsx(b"this is not a xlsx file")

    def test_non_wechat_xlsx_raises_valueerror(self):
        """非微信账单的 xlsx 抛 ValueError"""
        wb = Workbook()
        ws = wb.active
        ws.append(["随便写的内容"])
        buf = io.BytesIO()
        wb.save(buf)

        with pytest.raises(ValueError, match="微信支付账单明细"):
            parse_wechat_xlsx(buf.getvalue())

    def test_empty_xlsx_raises_valueerror(self):
        """空工作表抛 ValueError"""
        wb = Workbook()
        ws = wb.active
        # 写入标识行但没有表头
        ws.append(["微信支付账单明细"])
        buf = io.BytesIO()
        wb.save(buf)

        with pytest.raises(ValueError, match="表头行"):
            parse_wechat_xlsx(buf.getvalue())

    def test_no_data_after_header(self):
        """有表头但无数据行 → 返回空列表"""
        content = _build_wechat_xlsx([])
        result = parse_wechat_xlsx(content)
        assert result == []

    def test_empty_row_skipped(self):
        """空行被跳过"""
        content = _build_wechat_xlsx([
            _make_row(txn_id="001"),
            [None, None, None, None, None, None, None, None, None, None, None],
            _make_row(txn_id="002"),
        ])
        result = parse_wechat_xlsx(content)
        assert len(result) == 2


# ──────────── 支付方式 ────────────


class TestPaymentMethod:

    def test_payment_method_preserved(self):
        """支付方式原样保留"""
        content = _build_wechat_xlsx([_make_row(payment="招商银行储蓄卡(3717)")])
        result = parse_wechat_xlsx(content)
        assert result[0]["payment_method"] == "招商银行储蓄卡(3717)"

    def test_payment_method_slash(self):
        """支付方式为 / 时原样保留"""
        content = _build_wechat_xlsx([_make_row(payment="/")])
        result = parse_wechat_xlsx(content)
        assert result[0]["payment_method"] == "/"
